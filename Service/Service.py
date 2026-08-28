from tkinter import messagebox
import subprocess
import openpyxl
from openpyxl import Workbook
import os
from pathlib import Path
import re

class Service :
    def __init__(self, root, ui, util):
        self.root = root
        self.ui = ui
        self.util = util

        # 엑셀파일 저장경로

        #self.excel_path = r"\\172.26.21.20\ESU\esu.xlsx" #인천공항 공유
        #self.excel_path = r"\\172.28.243.228\ESU\esu.xlsx" #내부테스트용
        self.excel_path = r"c:\ESU\esu.xlsx"

        # 인증 cdkey 값 입력
        self.cdkey = "PGD98-N4MWG-9YKMW-P83F4-PPVPJ"

        # 정품 인증 ID값 입력
        self.activation_id = "f520e45e-7413-4a34-a497-d2765967d094"
        self.current_ip = self.util.get_local_ip()
        self.current_version = self.util.get_windows_build()

    def ensure_valid_ip(self):
        if self.util.is_valid_local_ipv4(self.current_ip):
            return True

        log_path = self.util.get_log_path()
        messagebox.showerror(
            "IP 확인 오류",
            f"현재 PC의 유효한 IP를 가져오지 못했습니다.\n\n"
            f"현재 IP: {self.current_ip}\n"
            f"로그 파일: {log_path}\n\n"
            f"네트워크 연결, 방화벽, VPN, IP 할당 상태를 확인해주세요."
        )
        self.ui.status_label.config(text="IP 확인 실패", fg="red")
        return False

    # 엑셀 파일 및 폴더 확인/생성
    def ensure_excel_exists(self):

        folder = Path(r"c:\ESU")
        if not folder.exists():
         folder.mkdir(parents=True, exist_ok=True)
        
        try:
            if not os.path.exists(self.excel_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "ESU인증"
                ws['A1'] = 'IP'
                ws['B1'] = 'DTI'
                ws['C1'] = '확인'
                ws['D1'] = 'END'
                ws['E1'] = 'PRE'  # PRE 열 추가

                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 70
                ws.column_dimensions['C'].width = 70

                wb.save(self.excel_path)
            else:
                # 파일이 있으면 헤더 확인
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active
                if ws['D1'].value != 'END':
                    ws['D1'] = 'END'
                if ws['E1'].value != 'PRE':
                    ws['E1'] = 'PRE'
                wb.save(self.excel_path)
                wb.close()
        except PermissionError:
            messagebox.showerror("파일 오류", "엑셀 파일(esu.xlsx)이 열려있습니다.\n파일을 닫고 프로그램을 다시 실행해주세요.")
    
    # cmd명령어 실행
    def run_command(self, command):
        try:
            result = subprocess.run(
                command,
                shell=True,
                capture_output=True,
                text=True,
                encoding='cp949',
                errors='ignore'
            )
            return result.stdout + result.stderr
        except Exception as e:
            return f"오류: {str(e)}"
    
    # A작업: Windows 버전 확인 후 CD키 설치 후 설치ID(DTI)값 수집후 엑셀에 저장
    def check_installation_id(self):
        if not self.ensure_valid_ip():
            return

        # 먼저 Windows 버전 확인
        if not self.util.check_windows_version():
            # 버전이 맞지 않으면 PRE 열에 "설치 불가" 저장
            self.save_pre_status("설치 불가")
            self.util.write_log(
                "Windows 버전 불일치로 설치 불가 처리",
                f"ip: {self.current_ip}\ncurrent_version: {self.current_version}\nrequired_version: 19045.6456 또는 19045.6466",
            )
            messagebox.showerror(
                "버전 불일치", 
                f"Windows 10 버전이 요구사항과 맞지 않습니다.\n\n"
                f"현재 버전: {self.current_version}\n"
                f"필요 버전: 19045.6456\n\n"
                f"엑셀 PRE 열에 '설치 불가'로 기록되었습니다."
            )
            self.ui.status_label.config(text="버전 불일치 - 설치 불가", fg="red")
            return
        
        # 버전이 맞으면 CD 키 확인
        if not self.cdkey:
            messagebox.showwarning("입력 오류", "CD 키를 입력해주세요!")
            return
        
        self.ui.status_label.config(text="설치ID 수집 중...", fg="blue")
        self.root.update()
        
        # CD 키 설치
        cmd_ipk = f'cscript //Nologo c:\\windows\\system32\\slmgr.vbs /ipk {self.cdkey}'
        self.run_command(cmd_ipk)
        
        # DTI 값 가져오기
        cmd_dti = f'cscript //Nologo c:\\windows\\system32\\slmgr.vbs /dti {self.activation_id}'
        result_dti = self.run_command(cmd_dti)
        
        dti_value = self.extract_dti_value(result_dti)
        
        if dti_value:
            # 엑셀 저장 시도 (성공/실패 반환 받음)
            if self.save_to_excel(dti_value):
                messagebox.showinfo("성공", f"설치ID가 수집되었습니다.\n\nDTI: {dti_value}\n\n엑셀 파일에 저장되었습니다.")
                self.ui.status_label.config(text="설치 완료", fg="green")
            else:
                self.ui.status_label.config(text="저장 실패 (엑셀 열림)", fg="red")
        else:
            messagebox.showerror("오류", f"설치ID 수집 실패\n\n{result_dti}")
            self.ui.status_label.config(text="설치 실패", fg="red")
    
    # PRE 열에 상태 저장(윈도우 버전 불일치 시 사용)
    def save_pre_status(self, status):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            ip_found = False
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == self.current_ip:
                    ws.cell(row, 5, status)  # E열(5번째 열)에 저장
                    ip_found = True
                    break
            
            if not ip_found:
                new_row = ws.max_row + 1
                ws.cell(new_row, 1, self.current_ip)
                ws.cell(new_row, 5, status)
            
            wb.save(self.excel_path)
            wb.close()
            return True
            
        except PermissionError:
            self.util.write_log(
                "PRE 상태 저장 실패: 엑셀 파일 열림",
                f"ip: {self.current_ip}\nstatus: {status}\nexcel_path: {self.excel_path}",
            )
            messagebox.showerror("저장 오류", "엑셀 파일(esu.xlsx)이 열려있습니다.\n파일을 닫고 다시 시도해주세요.")
            return False
        except Exception as e:
            self.util.write_log(
                "PRE 상태 저장 중 예외 발생",
                f"ip: {self.current_ip}\nstatus: {status}\nexcel_path: {self.excel_path}\nerror: {type(e).__name__}: {e}",
            )
            messagebox.showerror("오류", f"엑셀 저장 중 오류 발생: {e}")
            return False
    
    # 설치ID(DTI) 값 추출
    def extract_dti_value(self, output):
        lines = output.strip().split('\n')
        for line in lines:
            if '설치 ID' in line or 'Installation ID' in line:
                parts = line.split(':')
                if len(parts) > 1:
                    return parts[1].strip()
        
        # 결과가 숫자만으로 길게 나오거나 하는 경우 그대로 반환
        clean_output = output.strip()
        if len(clean_output) > 40:
            return clean_output
            
        return clean_output
    
    # 엑셀에 IP와 DTI 저장
    def save_to_excel(self, dti_value):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            ip_found = False
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == self.current_ip:
                    ws.cell(row, 2, dti_value)
                    ip_found = True
                    break
            
            if not ip_found:
                new_row = ws.max_row + 1
                ws.cell(new_row, 1, self.current_ip)
                ws.cell(new_row, 2, dti_value)
            
            wb.save(self.excel_path)
            wb.close()
            return True
            
        except PermissionError:
            self.util.write_log(
                "DTI 저장 실패: 엑셀 파일 열림",
                f"ip: {self.current_ip}\nexcel_path: {self.excel_path}",
            )
            messagebox.showerror("저장 오류", "엑셀 파일(esu.xlsx)이 열려있습니다.\n파일을 닫고 다시 시도해주세요.")
            return False
        except Exception as e:
            self.util.write_log(
                "DTI 저장 중 예외 발생",
                f"ip: {self.current_ip}\nexcel_path: {self.excel_path}\nerror: {type(e).__name__}: {e}",
            )
            messagebox.showerror("오류", f"엑셀 저장 중 오류 발생: {e}")
            return False
    
    # B작업: 인증 및 라이선스 상태 확인
    def activate_esu(self):
        if not self.ensure_valid_ip():
            return

        self.ui.status_label.config(text="인증 진행 중...", fg="blue")
        self.root.update()
        
        # 엑셀에서 확인 값 가져오기
        confirm_value = self.get_confirm_value_from_excel()
        
        if confirm_value == "OPEN_ERROR":
            self.ui.status_label.config(text="엑셀 파일 열려있음", fg="red")
            return

        if not confirm_value:
            messagebox.showerror("오류", f"엑셀 파일에서 IP({self.current_ip})에 해당하는 확인 값을 찾을 수 없습니다.\n'확인' 항목에 값을 입력해주세요.")
            self.ui.status_label.config(text="인증 실패 (값 없음)", fg="red")
            return
        
        # 인증 실행
        cmd_atp = f'cscript //Nologo c:\\windows\\system32\\slmgr.vbs /atp {confirm_value} {self.activation_id}'
        result_atp = self.run_command(cmd_atp)
        
        # 라이선스 상태 확인
        self.ui.status_label.config(text="라이선스 상태 확인 중...", fg="blue")
        self.root.update()
        
        # cmd_dlv = 'cscript //Nologo c:\\windows\\system32\\slmgr.vbs /dlv'
        # result_dlv = self.run_command(cmd_dlv)
        
        if self.activation_id:
            cmd_dlv_esu = f'cscript //Nologo c:\\windows\\system32\\slmgr.vbs /dlv {self.activation_id}'
            result_dlv = self.run_command(cmd_dlv_esu)
        
        license_status = self.extract_license_status(result_dlv)
        
        if license_status == "사용 허가됨":
            # 엑셀에 라이선스 상태 저장
            if self.save_license_status_to_excel(license_status):
                messagebox.showinfo("성공", 
                    f"인증이 완료되었습니다!\n\n"
                    f"인증 결과: 인증성공!\n\n"
                    f"라이선스 상태: {license_status}\n\n"
                    f"엑셀 파일 END 열에 저장되었습니다.")
                self.ui.status_label.config(text="인증 및 상태 확인 완료", fg="green")
            else:
                self.ui.status_label.config(text="저장 실패 (엑셀 열림)", fg="red")
        else:
            messagebox.showwarning("결과", 
                f"인증 결과:\n{result_atp}\n\n"
                f"QDFWW 키의 라이선스 상태를 찾을 수 없습니다.")
            self.ui.status_label.config(text="인증 실패 (상태 미확인)", fg="orange")
    
    # ESU 라이선스 상태 추출
    def extract_license_status(self, dlv_output):
        block = self.extract_esu_block(dlv_output)
        if block:
            status = self.extract_license_status_from_block(block)
            if status:
                return status
        return self.extract_license_status_from_block(dlv_output)

    def extract_esu_activation_id(self, dlv_output):
        block = self.extract_esu_block(dlv_output)
        if not block:
            return None
        pattern_kor = r"정품 인증 ID\s*:\s*([^\r\n]+)"
        match = re.search(pattern_kor, block, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        pattern_eng = r"Activation ID\s*:\s*([^\r\n]+)"
        match_eng = re.search(pattern_eng, block, re.IGNORECASE)
        if match_eng:
            return match_eng.group(1).strip()
        return None

    def extract_esu_block(self, dlv_output):
        blocks = re.split(r"\r?\n\r?\n+", dlv_output)
        for block in blocks:
            if re.search(r"\bESU\b", block, re.IGNORECASE):
                return block
        for block in blocks:
            if re.search(r"부분 제품 키\s*:\s*QDFWW", block, re.IGNORECASE):
                return block
            if re.search(r"Partial Product Key\s*:\s*QDFWW", block, re.IGNORECASE):
                return block
        return None

    def extract_license_status_from_block(self, text):
        pattern_kor = r"라이선스 상태\s*:\s*([^\r\n]+)"
        match = re.search(pattern_kor, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        pattern_eng = r"License Status\s*:\s*([^\r\n]+)"
        match_eng = re.search(pattern_eng, text, re.IGNORECASE)
        if match_eng:
            return match_eng.group(1).strip()
        return None
    
    # 현재 PC IP에 해당하는 라이선스 상태를 엑셀에 저장
    def save_license_status_to_excel(self, license_status):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            ip_found = False
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == self.current_ip:
                    ws.cell(row, 4, license_status)  # D열(4번째 열)에 저장
                    ip_found = True
                    break
            
            if not ip_found:
                new_row = ws.max_row + 1
                ws.cell(new_row, 1, self.current_ip)
                ws.cell(new_row, 4, license_status)

            wb.save(self.excel_path)
            wb.close()
            return True
            
        except PermissionError:
            self.util.write_log(
                "라이선스 상태 저장 실패: 엑셀 파일 열림",
                f"ip: {self.current_ip}\nlicense_status: {license_status}\nexcel_path: {self.excel_path}",
            )
            messagebox.showerror("저장 오류", "엑셀 파일(esu.xlsx)이 열려있습니다.\n파일을 닫고 다시 시도해주세요.")
            return False
        except Exception as e:
            self.util.write_log(
                "라이선스 상태 저장 중 예외 발생",
                f"ip: {self.current_ip}\nlicense_status: {license_status}\nexcel_path: {self.excel_path}\nerror: {type(e).__name__}: {e}",
            )
            messagebox.showerror("오류", f"저장 중 오류 발생: {e}")
            return False
    
    # 현재 PC IP에 해당하는 확인 값 가져오기
    def get_confirm_value_from_excel(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            
            val = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 1).value == self.current_ip:
                    val = ws.cell(row, 3).value
                    break
            
            wb.close()
            if val:
                return str(val).strip()
            return None
            
        except PermissionError:
            self.util.write_log(
                "확인 값 읽기 실패: 엑셀 파일 열림",
                f"ip: {self.current_ip}\nexcel_path: {self.excel_path}",
            )
            messagebox.showerror("읽기 오류", "엑셀 파일(esu.xlsx)이 열려있습니다.\n파일을 닫고 다시 시도해주세요.")
            return "OPEN_ERROR"
        except Exception as e:
            self.util.write_log(
                "확인 값 읽기 중 예외 발생",
                f"ip: {self.current_ip}\nexcel_path: {self.excel_path}\nerror: {type(e).__name__}: {e}",
            )
            return None
